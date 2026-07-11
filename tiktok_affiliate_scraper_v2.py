# """
# Tujuan: Skrip Python untuk menarik data performa video dari profil TikTok publik.
# Caller: app/api/scrape/route.ts
# Dependensi: yt-dlp, pandas, openpyxl
# Main Functions: scrape_tiktok, custom_scrape_fallback
# Side Effects: Melakukan HTTP requests ke TikTok embed API, menulis file Excel jika tidak dipanggil dengan --json.
# """

import yt_dlp
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, timezone
import os, time, random, sys, json, argparse

# ============================================================
#  ⚙️  PENGATURAN — Edit bagian ini saja
# ============================================================
USER_DEFAULT     = "dutaparfumlokal"
HARI_DEFAULT     = 1
BATAS_DEFAULT    = 400
DELAY_MIN        = 2.0
DELAY_MAX        = 4.5
# ============================================================

KAMUS_HARI = {
    'Monday': 'Senin', 'Tuesday': 'Selasa', 'Wednesday': 'Rabu',
    'Thursday': 'Kamis', 'Friday': 'Jumat', 'Saturday': 'Sabtu', 'Sunday': 'Minggu'
}
WIB = timezone(timedelta(hours=7))  # UTC+7
GREY   = "F5F5F5"
DARK   = "4A235A"

json_mode = False

def log_print(*args, **kwargs):
    if json_mode:
        print(*args, file=sys.stderr, **kwargs)
    else:
        print(*args, **kwargs)

def hitung_er(views, likes, komentar, share):
    if not views: return 0.0
    return round(((likes or 0) + (komentar or 0) + (share or 0)) / views * 100, 2)

def _border():
    t = Side(style='thin', color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)

def _header_row(ws, row_num, fill_color, font_color="FFFFFF"):
    for cell in ws[row_num]:
        cell.font      = Font(bold=True, color=font_color, name='Arial', size=10)
        cell.fill      = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = _border()

def _autofit(ws, min_w=10, max_w=40):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, min_w), max_w)


# ─── SCRAPER ────────────────────────────────────────────────

# ─── SCRAPER ────────────────────────────────────────────────

def custom_scrape_fallback(username, rentang_hari, batas):
    log_print("\n⚠️ Metode standar yt-dlp gagal. Mengaktifkan fallback Scraper Embed lokal...")
    data = []
    
    # 1. Temukan secUid dari profil / embed video pertama
    browsers = ['chrome', 'safari', 'firefox', 'opera', 'edge', None]
    sec_uid = None
    ydl = None
    used_browser = None
    
    # Coba temukan browser yang cookies-nya bisa memuat halaman embed profil
    for browser in browsers:
        try:
            opts = {
                'quiet': True,
                'ignoreerrors': True,
            }
            if browser:
                opts['cookiesfrombrowser'] = browser
            
            temp_ydl = yt_dlp.YoutubeDL(opts)
            embed_url = f"https://www.tiktok.com/embed/@{username}"
            req = temp_ydl.urlopen(embed_url)
            html = req.read().decode('utf-8')
            
            if html and len(html) > 10000:
                import re
                # Cari satu video ID
                video_ids = list(set(re.findall(r'/video/(\d+)', html)))
                if video_ids:
                    # Fetch video embed untuk dapat secUid
                    time.sleep(1)
                    v_url = f"https://www.tiktok.com/embed/v2/{video_ids[0]}"
                    v_req = temp_ydl.urlopen(v_url)
                    v_html = v_req.read().decode('utf-8')
                    
                    state_match = re.search(r'id="__FRONTITY_CONNECT_STATE__"[^>]*>(.*?)</script>', v_html, re.DOTALL)
                    if state_match:
                        state_data = json.loads(state_match.group(1).strip())
                        embed_path = f"/embed/v2/{video_ids[0]}"
                        data_obj = state_data.get("source", {}).get("data", {}).get(embed_path, {})
                        sec_uid = data_obj.get("videoData", {}).get("authorInfos", {}).get("secUid")
                        
                        if sec_uid:
                            ydl = temp_ydl
                            used_browser = browser
                            break
        except Exception:
            continue
            
    if not sec_uid or not ydl:
        log_print("❌ Gagal mematangkan sesi scraper (secUid tidak ditemukan).")
        return data
        
    log_print(f"🔑 secUid ditemukan: {sec_uid}")
    log_print(f"🔑 Menggunakan cookies peramban: {used_browser or 'Tanpa Cookies'}")
    
    # 2. Ambil seluruh daftaran video ID menggunakan flat playlist tiktokuser:sec_uid
    try:
        flat_opts = {
            'quiet': True,
            'ignoreerrors': True,
            'extract_flat': True,
            'playlist_items': f"1-{batas}",
        }
        if used_browser:
            flat_opts['cookiesfrombrowser'] = used_browser
            
        with yt_dlp.YoutubeDL(flat_opts) as playlist_ydl:
            playlist_url = f"tiktokuser:{sec_uid}"
            playlist_info = playlist_ydl.extract_info(playlist_url, download=False)
            
            video_ids = []
            if playlist_info and 'entries' in playlist_info:
                for entry in playlist_info['entries']:
                    if entry and entry.get('id'):
                        video_ids.append(entry.get('id'))
                        
            log_print(f"📌 Berhasil mendeteksi {len(video_ids)} video dari katalog lengkap.")
            
            if not video_ids:
                return data
                
            # Cutoff mulai dari jam 00:00 tanggal terlama yang diminta (calendar days)
            batas_waktu = datetime.now(tz=WIB).replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=rentang_hari - 1)
            
            for idx, vid_id in enumerate(video_ids, 1):
                try:
                    time.sleep(round(random.uniform(DELAY_MIN, DELAY_MAX), 2))
                    video_url = f"https://www.tiktok.com/embed/v2/{vid_id}"
                    v_req = ydl.urlopen(video_url)
                    v_html = v_req.read().decode('utf-8')
                    
                    state_match = re.search(r'id="__FRONTITY_CONNECT_STATE__"[^>]*>(.*?)</script>', v_html, re.DOTALL)
                    if not state_match:
                        continue
                        
                    state_data = json.loads(state_match.group(1).strip())
                    embed_path = f"/embed/v2/{vid_id}"
                    data_obj = state_data.get("source", {}).get("data", {}).get(embed_path, {})
                    video_data = data_obj.get("videoData", {})
                    
                    if not video_data:
                        continue
                        
                    item_infos = video_data.get("itemInfos", {})
                    
                    # Parse waktu pembuatan
                    create_time_str = item_infos.get("createTime")
                    tanggal = "Tidak terlacak"
                    jam = "Tidak terlacak"
                    hari = "Tidak terlacak"
                    
                    if create_time_str:
                        ts = int(create_time_str)
                        dt = datetime.fromtimestamp(ts, tz=timezone.utc).astimezone(WIB)
                        if dt < batas_waktu:
                            # Hentikan request video berikutnya karena sudah di luar rentang hari (descending order)
                            # KECUALI jika berada di 3 video pertama (kemungkinan Pinned Video yang tanggalnya lama)
                            if idx <= 3:
                                log_print(f"  📌 #{idx} tgl {dt.strftime('%d/%m/%Y')} — Terdeteksi di luar rentang tapi dilewati karena berada di urutan awal (kemungkinan Pinned Video).")
                                continue
                            else:
                                log_print(f"\n⛔ #{idx} tgl {dt.strftime('%d/%m/%Y')} — di luar rentang. Berhenti.")
                                break
                        tanggal = dt.strftime('%d/%m/%Y')
                        jam     = dt.strftime('%H:%M')
                        hari    = KAMUS_HARI.get(dt.strftime('%A'), dt.strftime('%A'))
                    
                    views  = int(item_infos.get("playCount") or 0)
                    likes  = int(item_infos.get("diggCount") or 0)
                    komen  = int(item_infos.get("commentCount") or 0)
                    share  = int(item_infos.get("shareCount") or 0)
                    durasi = int(item_infos.get("video", {}).get("videoMeta", {}).get("duration") or 0)
                    desk   = item_infos.get("text") or ""
                    link   = f"https://www.tiktok.com/@{username}/video/{vid_id}"
                    er     = hitung_er(views, likes, komen, share)
                    
                    log_print(f"  #{idx:>3} | {tanggal} {jam} | Views: {views:>8,} | ER: {er:>5.1f}%")
                    
                    data.append({
                        'No'                 : idx,
                        'Hari'               : hari,
                        'Tanggal Upload'     : tanggal,
                        'Jam Upload'         : jam,
                        'Deskripsi'          : desk,
                        'Durasi (detik)'     : durasi,
                        'Views'              : views,
                        'Likes'              : likes,
                        'Komentar'           : komen,
                        'Share'              : share,
                        'Konten ID'          : vid_id,
                        'Link Video'         : link,
                    })
                except Exception as inner_e:
                    log_print(f"⚠️ Gagal memproses video {vid_id}: {inner_e}")
    except Exception as e:
        log_print(f"⚠️ Gagal menjalankan fallback scraper: {e}")
        
    return data

def scrape_tiktok(username, rentang_hari, batas):
    # Cutoff mulai dari jam 00:00 tanggal terlama yang diminta (calendar days)
    batas_waktu = datetime.now(tz=WIB).replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=rentang_hari - 1)
    url = f"https://www.tiktok.com/@{username}"
    log_print(f"\n{'='*60}\n  🤖 TikTok Affiliate Scraper — @{username}")
    log_print(f"  📅 Video sejak: {batas_waktu.strftime('%d/%m/%Y')}\n{'='*60}\n")

    browsers = ['chrome', 'safari', 'firefox', 'opera', 'edge', None]
    data = []
    
    for browser in browsers:
        ydl_opts = {
            'extract_flat': False, 
            'playlist_items': f'1-{batas}',
            'quiet': True, 
            'ignoreerrors': True, 
            'skip_download': True,
        }
        if browser:
            ydl_opts['cookiesfrombrowser'] = browser
            
        try:
            with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                info = ydl.extract_info(url, download=False, process=False)
                if not info or 'entries' not in info:
                    continue

                for i, entry in enumerate(info['entries'], 1):
                    if not entry: continue
                    time.sleep(round(random.uniform(DELAY_MIN, DELAY_MAX), 2))
                    try:
                        vid = ydl.process_ie_result(entry, download=False)
                    except Exception:
                        continue
                    if not vid: continue

                    ts  = vid.get('timestamp')
                    tgl = vid.get('upload_date')
                    tanggal = jam = hari = "Tidak terlacak"

                    if ts:
                        dt = datetime.fromtimestamp(ts, tz=timezone.utc).astimezone(WIB)
                        if dt < batas_waktu:
                            # Hentikan request video berikutnya karena sudah di luar rentang hari (descending order)
                            # KECUALI jika berada di 3 video pertama (kemungkinan Pinned Video yang tanggalnya lama)
                            if i <= 3:
                                log_print(f"  📌 #{i} tgl {dt.strftime('%d/%m/%Y')} — Terdeteksi di luar rentang tapi dilewati karena berada di urutan awal (kemungkinan Pinned Video).")
                                continue
                            else:
                                log_print(f"\n⛔ #{i} tgl {dt.strftime('%d/%m/%Y')} — di luar rentang. Berhenti.")
                                break
                        tanggal = dt.strftime('%d/%m/%Y')
                        jam     = dt.strftime('%H:%M')
                        hari    = KAMUS_HARI.get(dt.strftime('%A'), dt.strftime('%A'))
                    elif tgl:
                        tanggal = f"{tgl[6:8]}/{tgl[4:6]}/{tgl[0:4]}"

                    vid_id = vid.get('id', '')
                    link   = vid.get('webpage_url') or vid.get('url', '')
                    if vid_id and 'tiktok.com' not in link:
                        link = f"https://www.tiktok.com/@{username}/video/{vid_id}"

                    views  = vid.get('view_count') or 0
                    likes  = vid.get('like_count') or 0
                    komen  = vid.get('comment_count') or 0
                    share  = vid.get('repost_count') or vid.get('share_count') or 0
                    durasi = vid.get('duration') or 0
                    desk   = vid.get('description') or vid.get('title') or ''
                    er     = hitung_er(views, likes, komen, share)

                    log_print(f"  #{i:>3} | {tanggal} {jam} | Views: {views:>8,} | ER: {er:>5.1f}%")

                    data.append({
                        'No'                 : i,
                        'Hari'               : hari,
                        'Tanggal Upload'     : tanggal,
                        'Jam Upload'         : jam,
                        'Deskripsi'          : desk,
                        'Durasi (detik)'     : durasi,
                        'Views'              : views,
                        'Likes'              : likes,
                        'Komentar'           : komen,
                        'Share'              : share,
                        'Konten ID'          : vid_id,
                        'Link Video'         : link,
                    })
                
                if data:
                    break
        except Exception:
            continue
        
    if not data:
        # Jika hasil kosong, panggil fallback
        return custom_scrape_fallback(username, rentang_hari, batas)
        
    return data


# ─── EXCEL BUILDER ──────────────────────────────────────────

def buat_excel(data, username, rentang_hari):
    if not data:
        log_print("\n⚠️  Tidak ada data.")
        return

    df = pd.DataFrame(data)
    nama_file = f"{username}_{rentang_hari}hari_affiliate.xlsx"

    with pd.ExcelWriter(nama_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Data Video', index=False)

    wb = load_workbook(nama_file)
    ws = wb['Data Video']
    _style_data(ws, df)
    wb.save(nama_file)

    log_print(f"\n{'='*60}")
    log_print(f"  ✅ File tersimpan: {nama_file}")
    log_print(f"  📊 Total video : {len(df)}")
    log_print(f"\n  📝 Langkah selanjutnya:")
    log_print(f"{'='*60}\n")
    return nama_file


# ─── STYLING ────────────────────────────────────────────────

def _style_data(ws, df):
    ws.row_dimensions[1].height = 32
    _header_row(ws, 1, DARK)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        bg = "FFFFFF" if row_idx % 2 == 0 else GREY
        for cell in row:
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.border    = _border()
            cell.font      = Font(name='Arial', size=9)
            cell.alignment = Alignment(vertical='center', wrap_text=False)

    _autofit(ws)


# ─── MAIN ───────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="TikTok Affiliate Scraper")
    parser.add_argument("--username", type=str, default=USER_DEFAULT, help="Username TikTok")
    parser.add_argument("--days", type=int, default=HARI_DEFAULT, help="Rentang hari")
    parser.add_argument("--batas", type=int, default=BATAS_DEFAULT, help="Batas video")
    parser.add_argument("--json", action="store_true", help="Output as JSON directly to stdout")

    args = parser.parse_args()
    
    if args.json:
        json_mode = True

    data = scrape_tiktok(args.username, args.days, args.batas)
    
    if args.json:
        print(json.dumps(data))
    else:
        if data:
            buat_excel(data, args.username, args.days)
        else:
            print("Tidak ada data yang berhasil diambil.")